# -*- coding: utf-8 -*-
import openpyxl as opx
import xlsxwriter
from configparser import ConfigParser
import re
import subprocess, glob, os, shutil
from time import time
from datetime import datetime

from smtplib import SMTP
from email.mime.multipart import MIMEMultipart
#from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

#File names
outputFile = "items_out.xlsx"
workDir = os.path.dirname(os.path.abspath(__file__))
dstDir = ''
eMails = ['d.chestnov@inlinegroup.ru', 'a.korolkov@inlinegroup.ru']


#Define column names and their indexes
PN = 2
Description = 3
vendorName = 5
sTip = 6
Category = 7
Subcat1 = 8
Subcat2 = 9
Subcat3 = 10

LOGGING = "YES"
logFile = 'processing.log'

# List of vendors in lowercase
vendor_list = [ 'cisco', 'huawei', 'checkpoint', 'linksys', 'paloalto', 
				'nutanix', 'mellanox', 'juniper', 'sterra', 'unify', 
				'crestron', 'aruba', 'avaya'
				]


vendor_dict = { 'sterra':		['ерра'],
				'checkpoint':	['check'],
				'cisco':		['cisco', 'linksys'],
				'paloalto':		['palo'],
				'huawei':		['huawei'],
				'nutanix':		['nutanix'],
				'juniper':		['juniper'],
				'mellanox':		['mellanox'],
				'unify':		['unify'],
				'avaya':		['avaya'],
				'crestron':		['crestron'],
				'aruba':		['aruba']
				}




# Define columns to fill
editCols = [sTip, Category, Subcat1, Subcat2, Subcat3]

#email sender
def sendmail(sender, recipients):
    mail_content =  '''
                        Hello,
                        This is a test mail.
                        In this mail we are sending some attachments.
                        The mail is sent using Python SMTP library.
                        Thank You
                    '''

    server = '10.8.50.75'

    message = MIMEMultipart()
    message['From'] = sender
    message['To'] = ", ".join(recipients)
    message['Subject'] = 'Обработанная номенклатура для загрузки'

    payload = MIMEBase('application', 'octate-stream')
    payload.set_payload(open(outputFile, 'rb').read())
    encoders.encode_base64(payload) 
    payload.add_header('Content-Disposition', 'attachment; filename = "items_out.xls"')
    message.attach(payload)

    session = SMTP(server) 
    #session.starttls() 
    #session.login(sender, password) 
    text = message.as_string()
    session.sendmail(sender, recipients, text)
    session.quit()
    print('Mail Sent')


# Get data from Excel
class xlsReader:
    def __init__(self, inFile):
        self.WB = opx.load_workbook(inFile)
        self.WS = self.WB.worksheets[0]

        
    def get_header(self):
        firstRow = next(self.WS.iter_rows(min_row=1, max_row=1)) 
        return [c.value for c in firstRow]


    def get_row(self, rowIndex):        
        row = self.WS[rowIndex]
        return [c.value for c in row ]

    
    def get_max_row(self):
        return self.WS.max_row
    
    
# Write data to Excel
class xlsWriter:
    def __init__(self, outFile):
        self.WB = xlsxwriter.Workbook(outFile)
        self.WS = self.WB.add_worksheet()
        self.file = outFile

                
    def save_wb(self):
        self.WB.save(self.file)

                   
    def write_row(self, rowIndex, rowList):
        for colIndex, item in enumerate(rowList):
             self.WS.write(rowIndex, colIndex, item)

             
# Read RegEx from config file to Dictionary
class MyParser(ConfigParser):        
    def as_dict(self):
        d = dict(self._sections)
        for k in d:
            d[k] = dict(self._defaults, **d[k])
            d[k].pop('__name__', None)
        for k, v in d.items():
            for i, j in v.items():
                d[k][i] = "|".join(j.split("\n"))
        return d


# General Vendor definition
class Vendor:
    def __init__(self, name, row):
        self.name = name        
        self.cfg = MyParser()
        self.cfg.optionxform = lambda option: option
        self.cfg.read(self.name + '.conf')
        self.regex = self.cfg.as_dict()
        self.index = list(self.regex.keys())
        self.row = row
        if name == 'cisco':
            self.checkCols = [PN, Description]
        else:
            self.checkCols = [Description]

        
    def __check_cell(self, col):
        cellValue = self.row[col]
        rez = []
        for c, r in zip(editCols, self.index): 
            reg = self.regex[r]
            z = False
            for k, v in reg.items():
                if re.search(v, cellValue):
                    z = k
                    break
            rez.append(z)
        return rez

                
    def get_rezult(self):
        rez = []
        for col in self.checkCols:
            rez.append(self.__check_cell(col))
        row = rez[0]
        if len(self.checkCols) > 1:
            for i, value in enumerate(row):
                if not value:
                    row[i] = rez[1][i]
        for i in range(len(row)):
            if not row[i]:
                row[i] = 'Прочее'        
        self.row[sTip:] = row
        #if self.name == 'checkpoint':
        #    row[vendorName] = 'CheckPoint'
        print (self.row)
        return self.row
                

def correct_vn(vName):
	vn = vName.lower().replace(' ', '')
	for k, v in vendor_dict.items():
		for i in v:
			if i in vn:
				return k
 
    
def get_latest_file():
    subprocess.call(r'net use z: /del /Y', shell=True)
    subprocess.call(r'net use z: \\inlinegroup.ru\Public\ППВВ\Треш /user:d.chestnov@inlinegroup.ru Gsm6raiN', shell=True)
    os.chdir(r'Z:')
    list_of_files = glob.glob("*.xlsx") 
    latest_file = max(list_of_files, key=os.path.getctime)   
    timeDelta = time() - 3600*24*10  #Check only files not older than 6 days
    fileCreation = os.path.getctime(latest_file) 
    if latest_file and (fileCreation > timeDelta):
        print ('\nProcessing file: ', latest_file)
        shutil.copy2(latest_file, workDir)
        os.chdir(workDir)
        subprocess.call(r'net use z: /del /Y', shell=True)
        return latest_file
    else:
        print ('No new file to process. Exit')
        subprocess.call(r'net use z: /del /Y', shell=True) 
        exit() 


def logging(info):
	time = "[%s] " %datetime.strftime(datetime.now(), "%d/%m/%Y %H:%M")
	logStr = time + '\t\t' + info + '\n'
	print (logStr)
	if LOGGING == "YES":
		with open(logFile, 'a') as f:
			f.write(logStr)    
           

def copy_file(f):
	subprocess.call(r'net use z: /del /Y', shell=True)
	subprocess.call(r'net use z: \\inlinegroup.ru\Public\ППВВ\Треш\Upload /user:d.chestnov@inlinegroup.ru Gsm6raiN', shell=True)
	shutil.copy2(f, r'Z:')
	subprocess.call(r'net use z: /del /Y', shell=True) 


def main():

    inputFile = get_latest_file()
    xr = xlsReader(inputFile) 
    xw = xlsWriter(outputFile)
    header = xr.get_header()
    xw.write_row(0, header)    
    last_row = xr.get_max_row()
    k = 1
    for index in range(1, last_row):
        row = xr.get_row(index)
        vnRaw = row[vendorName]
        if vnRaw:
            vn = correct_vn(vnRaw)
            if vn:
                vendor = Vendor(vn, row)
                newRow = vendor.get_rezult()
                xw.write_row(k, newRow)
                k += 1

    xw.save_wb()
    print ('\n [+] Обработка завершена...')
    logging (inputFile)
    copy_file(outputFile)
    #sendmail('d.chestnov@inlinegroup.ru', eMails)
    

         

if __name__ == '__main__':
	main()

